import os
from collections import defaultdict
from lxml import etree
import decimal
import openpyxl
import datetime


def clear_tag(tag_str: str) -> str:
    return tag_str[tag_str.find('}') + 1:]



# https://stackoverflow.com/questions/7684333/converting-xml-to-dictionary-using-elementtree
def etree_to_dict(t):
    d = {clear_tag(t.tag): {} if t.attrib else None}
    children = list(t)
    if children:
        dd = defaultdict(list)
        for dc in map(etree_to_dict, children):
            for k, v in dc.items():
                dd[k].append(v)
        d = {clear_tag(t.tag): {clear_tag(k): v[0] if len(v) == 1 else v
                                for k, v in dd.items()}}
    if t.attrib:
        d[t.tag].update(('@' + k, v)
                        for k, v in t.attrib.items())
    if t.text:
        text = t.text.strip()
        if children or t.attrib:
            if text:
                d[t.tag]['#text'] = text
        else:
            d[t.tag] = text
    return d


def main():
    xml_filenames = []

    for root, dirs, files in os.walk('XML'):
        for file in files:
            if file.endswith('.xml'):
                xml_filenames.append(os.path.join(root, file))

    xml_stat = dict()

    d1 = datetime.datetime(2021, 1, 1)
    d2 = datetime.datetime(2022, 9, 12)

    time_stamp = datetime.datetime.now()

    with open('stat.log', 'w', encoding='utf-8') as log_out:
        log_out.write(f'Начало работы программы: {time_stamp}\n')

    for xml_file in xml_filenames:
        try:
            xml_root = etree.parse(xml_file).getroot()
            xml_dict = etree_to_dict(xml_root)

            facts_list = xml_dict['data']['package']['elements']['fact']

            if isinstance(facts_list, dict):
                facts_list = [facts_list]

            for fact in facts_list:

                lmsz_id = fact['LMSZID']

                decision_date = datetime.datetime.strptime(fact['decision_date'], '%Y-%m-%d')

                if not(d1 <= decision_date <= d2):
                    continue

                decision_year = decision_date.year

                fact_amount = 0
                if 'monetary_form' in fact['assignment_info'].keys():
                    fact_amount = decimal.Decimal(fact['assignment_info']['monetary_form']['amount'])
                elif 'natural_form' in fact['assignment_info'].keys():
                    fact_amount = decimal.Decimal(fact['assignment_info']['natural_form']['equivalentAmount'])

                if f'{decision_year}+{lmsz_id}' not in xml_stat.keys():
                    xml_stat[f'{decision_year}+{lmsz_id}'] = {'facts_count': 0, 'total_amount': 0}

                xml_stat[f'{decision_year}+{lmsz_id}']['facts_count'] += 1
                xml_stat[f'{decision_year}+{lmsz_id}']['total_amount'] += fact_amount

        except Exception as e:
            with open('stat.log', 'a', encoding='utf-8') as log_out:
                log_out.write(f'{xml_file}\n')
                log_out.write(f'{e}\n')
            continue

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'инфо'
    ws['A1'] = 'Отметка времени:'
    ws['A2'] = time_stamp
    ws['A4'] = 'Файлов обработано:'
    ws['A5'] = len(xml_filenames)
    ws['A7'] = 'Список обработанных файлов:'
    for i, file_name in enumerate(xml_filenames, 8):
        ws[f'A{i}'] = file_name.lstrip('XML\\')

    ws = wb.create_sheet('статистика')
    ws['A1'] = 'год'
    ws['B1'] = 'ИД меры'
    ws['C1'] = 'код меры'
    ws['D1'] = 'кол-во фактов'
    ws['E1'] = 'общая сумма'

    for i, k in enumerate(sorted(xml_stat.keys()), 2):
        ws[f'A{i}'], ws[f'B{i}'] = k.split('+')
        ws[f'D{i}'] = xml_stat[k]['facts_count']
        ws[f'E{i}'] = xml_stat[k]['total_amount']

    wb.save(f'xml_stats_{time_stamp.strftime("%d_%m_%Y_%H%M")}.xlsx')

    with open('stat.log', 'a', encoding='utf-8') as log_out:
        log_out.write(f'Конец работы программы: {datetime.datetime.now()}\n')


if __name__ == '__main__':
    main()
